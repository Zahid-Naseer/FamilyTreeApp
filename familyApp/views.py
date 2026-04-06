import random, string
import datetime
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Prefetch
from django.shortcuts import render, redirect, get_object_or_404
from .models import Family, Person, Marriage
from django.contrib.auth import update_session_auth_hash
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ── HELPERS ──────────────────────────────────────────────────────

def make_invite_code():
    while True:
        code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        if not Family.objects.filter(invite_code=code).exists():
            return code


def get_user_family(user):
    return user.families.first()

def is_family_admin(user, family):
    return family.created_by == user

def can_edit_person(user, person, family):
    if person.claimed_by == user:
        return True  
    if person.claimed_by is None and is_family_admin(user, family):
        return True
    return False

def can_delete_person(user, person, family):
    # Admin can delete unclaimed persons
    if is_family_admin(user, family) and person.claimed_by is None:
        return True
    # The person who claimed the profile can delete their own
    if person.claimed_by == user:
        return True
    return False

# ── AUTH ─────────────────────────────────────────────────────────

def landing(request):
    if request.user.is_authenticated:
        return redirect('base')
    return render(request, 'landing.html')

def register_view(request):
    error = None
    if request.method == 'POST':
        action      = request.POST.get('action')
        username    = request.POST.get('username', '').strip()
        password    = request.POST.get('password', '')
        family_name = request.POST.get('family_name', '').strip()
        invite_code = request.POST.get('invite_code', '').strip().upper()

        if User.objects.filter(username=username).exists():
            error = "Username already taken."
        else:
            user = User.objects.create_user(username=username, password=password)
            if action == 'create':
                family = Family.objects.create(
                    name=family_name, created_by=user,
                    invite_code=make_invite_code()
                )
                family.members.add(user)
                login(request, user)
                return redirect('base')
            elif action == 'join':
                try:
                    family = Family.objects.get(invite_code=invite_code)
                    family.members.add(user)
                    login(request, user)
                    return redirect('base')
                except Family.DoesNotExist:
                    user.delete()
                    error = "Invalid invite code."
    return render(request, 'register.html', {'error': error})

def login_view(request):
    error = None
    if request.method == 'POST':
        user = authenticate(request,
            username=request.POST.get('username'),
            password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('base')
        error = "Invalid username or password."
    return render(request, 'login.html', {'error': error})

def logout_view(request):
    logout(request)
    return redirect('login')


# ── HOME ─────────────────────────────────────────────────────────

@login_required
def base(request):
    family = get_user_family(request.user)
    if not family:
        return redirect('landing')
    return render(request, 'base.html', {
        'family': family,
        'is_admin': is_family_admin(request.user, family),
        'invite_code': family.invite_code,
    })


# ── PUBLIC PROFILE ────────────────────────────────────────────────

@login_required
def account_settings(request):
    success = None
    error = None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_info':
            username = request.POST.get('username', '').strip()
            email    = request.POST.get('email', '').strip()

            if username and username != request.user.username:
                if User.objects.filter(username=username).exclude(pk=request.user.pk).exists():
                    error = "That username is already taken."
                else:
                    request.user.username = username

            request.user.email = email
            request.user.save()
            success = "Details updated successfully."

        elif action == 'change_password':
            current  = request.POST.get('current_password')
            new_pass = request.POST.get('new_password')
            confirm  = request.POST.get('confirm_password')

            if not request.user.check_password(current):
                error = "Current password is incorrect."
            elif new_pass != confirm:
                error = "New passwords do not match."
            elif len(new_pass) < 6:
                error = "Password must be at least 6 characters."
            else:
                request.user.set_password(new_pass)
                request.user.save()
                update_session_auth_hash(request, request.user)  # keeps user logged in
                success = "Password changed successfully."

    return render(request, 'account_settings.html', {
        'success': success,
        'error':   error,
    })
# ── TREE ─────────────────────────────────────────────────────────

def build_tree(person, visited=None):
    if visited is None:
        visited = set()
    if person.id in visited:
        return None
    visited.add(person.id)

    marriages = Marriage.objects.filter(
        Q(spouse1=person) | Q(spouse2=person)
    ).select_related('spouse1', 'spouse2')

    spouses = [
        m.spouse2 if m.spouse1_id == person.id else m.spouse1
        for m in marriages
    ]

    # Children where THIS person is the father
    children_as_father = Person.objects.filter(father=person).select_related('father', 'mother', 'claimed_by')

    # Children where THIS person is the mother AND
    # the father joined by marriage (so he has no birth subtree of his own)
    children_as_mother = Person.objects.filter(
        mother=person,
        father__joined_by_marriage=True
    ).select_related('father', 'mother', 'claimed_by')

    # Also children where this person is mother and father is NULL
    children_no_father = Person.objects.filter(
        mother=person,
        father__isnull=True
    ).select_related('father', 'mother', 'claimed_by')

    # Combine and deduplicate
    all_children = (children_as_father | children_as_mother | children_no_father).distinct()

    return {
        'person':   person,
        'spouses':  spouses,
        'children': [build_tree(c, visited) for c in all_children],
    }

@login_required
def tree_view(request):
    family = get_user_family(request.user)
    roots = Person.objects.filter(
        family=family, father__isnull=True,
        mother__isnull=True, joined_by_marriage=False , gender='M'
    ).select_related('father', 'mother', 'claimed_by')
    return render(request, 'treeView.html', {'trees': [build_tree(r) for r in roots]})


# ── PERSON DETAIL ─────────────────────────────────────────────────

@login_required
def person_detail(request, person_id):
    family = get_user_family(request.user)
    person = get_object_or_404(Person, pk=person_id, family=family)

    # Efficiently fetch all ancestors in one query
    ancestors = []
    current = person.father
    ancestors_set = set()
    while current:
        if current.id in ancestors_set:
            break  # Prevent infinite loops
        ancestors.append(current)
        ancestors_set.add(current.id)
        current = current.father
    ancestors.reverse()

    marriages = Marriage.objects.filter(
        Q(spouse1=person) | Q(spouse2=person)
    ).select_related('spouse1', 'spouse2')

    children = Person.objects.filter(Q(father=person) | Q(mother=person)).distinct()

    siblings = Person.objects.none()
    if person.father:
        siblings = siblings | Person.objects.filter(father=person.father)
    if person.mother:
        siblings = siblings | Person.objects.filter(mother=person.mother)
    siblings = siblings.exclude(pk=person.pk).distinct()

    already_claimed_other = (
    hasattr(request.user, 'person_profile')
    and request.user.person_profile != person
)

    return render(request, 'person_detail.html', {
    'person':     person,
    'ancestors':  ancestors,
    'marriages':  marriages,
    'children':   children,
    'siblings':   siblings,
    'is_admin':   is_family_admin(request.user, family),
    'can_edit':   can_edit_person(request.user, person, family),
    'can_delete': can_delete_person(request.user, person, family),
    'can_claim':  person.claimed_by is None,  # just this one check
})


# ── ADD PERSON ────────────────────────────────────────────────────

@login_required
def add_new_person(request):
    family  = get_user_family(request.user)
    males   = Person.objects.filter(family=family, gender='M')
    females = Person.objects.filter(family=family, gender='F')

    if request.method == 'POST':
        name          = request.POST.get('name', '').strip()
        gender        = request.POST.get('gender')
        father_id     = request.POST.get('father')
        mother_id     = request.POST.get('mother')
        mother_name   = request.POST.get('mother_name', '').strip()
        mother_birth  = request.POST.get('mother_birth_date')
        mother_bio    = request.POST.get('mother_bio', '')
        marriage_date = request.POST.get('marriage_date')
        father_name   = request.POST.get('father_name', '').strip()
        father_birth  = request.POST.get('father_birth_date')
        father_bio    = request.POST.get('father_bio', '')
        birth_date    = request.POST.get('birth_date')
        bio           = request.POST.get('bio', '')
        joined        = request.POST.get('joined_by_marriage') == 'on'
        is_me         = request.POST.get('is_me') == 'on'

        father = get_object_or_404(Person, pk=father_id, family=family) if father_id else None
        mother = get_object_or_404(Person, pk=mother_id, family=family) if mother_id else None

        if father and father.gender != 'M':
            return render(request, 'addNew.html', {
                'males': males, 'females': females,
                'error': f"{father.name} is not male and cannot be set as father."
            })
        if mother and mother.gender != 'F':
            return render(request, 'addNew.html', {
                'males': males, 'females': females,
                'error': f"{mother.name} is not female and cannot be set as mother."
            })

        # Both exist — ensure they are married
        if father and mother:
            ensure_marriage(father, mother, marriage_date)

        # Mother has no husband — auto create father + marriage
        if mother and not father and father_name:
            father = Person.objects.create(
                family=family,
                name=father_name,
                gender='M',
                birth_date=father_birth or None,
                bio=father_bio,
                joined_by_marriage=True,
                added_by=request.user
            )
            ensure_marriage(father, mother, marriage_date)

        # Father has no wife — auto create mother + marriage
        if father and not mother and mother_name:
            mother = Person.objects.create(
                family=family,
                name=mother_name,
                gender='F',
                birth_date=mother_birth or None,
                bio=mother_bio,
                joined_by_marriage=True,
                added_by=request.user
            )
            ensure_marriage(father, mother, marriage_date)

        person = Person.objects.create(
            family=family, name=name, gender=gender,
            birth_date=birth_date or None, bio=bio,
            father=father, mother=mother,
            joined_by_marriage=joined,
            added_by=request.user,
            claimed_by=request.user if is_me else None
        )

        if person.gender == 'F' and not person.father and not person.mother:
            person.joined_by_marriage = True
            person.save()

        next_url = request.POST.get('next', '')
        if next_url == 'tree':
            return redirect('treeView')

        return render(request, 'addNew.html', {
            'success': True, 'person': person,
            'males': males, 'females': females,
        })

    return render(request, 'addNew.html', {'males': males, 'females': females})
 


# ── EDIT PERSON ───────────────────────────────────────────────────

def ensure_marriage(father, mother, marriage_date=None):
    """Create marriage between father and mother if it doesn't exist."""
    if father and mother:
        already = Marriage.objects.filter(
            Q(spouse1=father, spouse2=mother) |
            Q(spouse1=mother, spouse2=father)
        ).exists()
        if not already:
            Marriage.objects.create(
                spouse1=father,
                spouse2=mother,
                marriage_date=marriage_date or None
            )

@login_required
def edit_person(request, person_id):
    family = get_user_family(request.user)
    person = get_object_or_404(Person, pk=person_id, family=family)

    if not can_edit_person(request.user, person, family):
        return render(request, 'error.html', {
            'message': f"{person.name}'s profile is managed by them. You cannot edit it."
        })

    if request.method == 'POST':
        person.name       = request.POST.get('name', person.name).strip()
        person.bio        = request.POST.get('bio', person.bio)
        person.birth_date = request.POST.get('birth_date') or None

        father_id = request.POST.get('father')
        mother_id = request.POST.get('mother')
        person.father = get_object_or_404(Person, pk=father_id, family=family) if father_id else None
        person.mother = get_object_or_404(Person, pk=mother_id, family=family) if mother_id else None

        person.save()

        # Auto-marry father and mother if both present and not already married
        ensure_marriage(person.father, person.mother)

        return redirect('person_detail', person_id=person.pk)

    males   = Person.objects.filter(family=family, gender='M').exclude(pk=person.pk)
    females = Person.objects.filter(family=family, gender='F').exclude(pk=person.pk)

    return render(request, 'edit_person.html', {
        'person':  person,
        'males':   males,
        'females': females,
    })


# ── CLAIM PROFILE ─────────────────────────────────────────────────

@login_required
def claim_profile(request, person_id):
    family = get_user_family(request.user)
    person = get_object_or_404(Person, pk=person_id, family=family)

    if person.claimed_by and person.claimed_by != request.user:
        return render(request, 'error.html', {
            'message': "This profile is already claimed by another account."
        })

    if request.method == 'POST':
        person.claimed_by = request.user
        person.save()
        return redirect('person_detail', person_id=person.pk)

    return render(request, 'claim_profile.html', {'person': person})


# ── DELETE PERSON ─────────────────────────────────────────────────

@login_required
def delete_person(request, person_id):
    family = get_user_family(request.user)
    person = get_object_or_404(Person, pk=person_id, family=family)

    if not can_delete_person(request.user, person, family):
        return render(request, 'error.html', {
            'message': "You cannot delete this person."
        })

    if request.method == 'POST':
        child_action = request.POST.get('child_action', 'detach')

        children = Person.objects.filter(
            Q(father=person) | Q(mother=person)
        )

        if child_action == 'delete_all':
            # Recursively delete all descendants
            def delete_recursive(p):
                descendants = Person.objects.filter(
                    Q(father=p) | Q(mother=p)
                )
                for d in descendants:
                    delete_recursive(d)
                Marriage.objects.filter(
                    Q(spouse1=p) | Q(spouse2=p)
                ).delete()
                p.delete()

            for child in children:
                delete_recursive(child)

        else:
            # Just detach — children become roots
            children.filter(father=person).update(father=None)
            children.filter(mother=person).update(mother=None)

        Marriage.objects.filter(
            Q(spouse1=person) | Q(spouse2=person)
        ).delete()
        person.delete()
        return redirect('treeView')

    children = Person.objects.filter(
        Q(father=person) | Q(mother=person)
    ).distinct()

    return render(request, 'confirm_delete.html', {
        'person':   person,
        'children': children,
    })


# ── ADD RELATIONSHIP ──────────────────────────────────────────────

@login_required
def add_relationship(request):
    family  = get_user_family(request.user)
    males   = Person.objects.filter(family=family, gender='M')
    females = Person.objects.filter(family=family, gender='F')
    error   = None

    if request.method == 'POST':
        s1_id = request.POST.get('spouse1')
        s2_id = request.POST.get('spouse2')
        date  = request.POST.get('marriage_date')

        if s1_id == s2_id:
            error = "A person cannot marry themselves."
        else:
            s1 = get_object_or_404(Person, pk=s1_id, family=family)
            s2 = get_object_or_404(Person, pk=s2_id, family=family)

            if s1.gender == s2.gender:
                error = "Marriage must be between a male and a female."

            elif Marriage.objects.filter(Q(spouse1=s1)|Q(spouse2=s1)).exists():
                error = f"{s1.name} is already married."

            elif Marriage.objects.filter(Q(spouse1=s2)|Q(spouse2=s2)).exists():
                error = f"{s2.name} is already married."

            elif Marriage.objects.filter(Q(spouse1=s1,spouse2=s2)|Q(spouse1=s2,spouse2=s1)).exists():
                error = f"{s1.name} and {s2.name} are already married to each other."

            else:
    # ── Age validation ──────────────────────────────
                if date:
                    try:
                        marriage_date = datetime.date.fromisoformat(date)

                        for person in [s1, s2]:
                            if person.birth_date:
                                # Marriage before birth
                                if marriage_date < person.birth_date:
                                    error = (
                                        f"Marriage date ({marriage_date.strftime('%d %b %Y')}) "
                                        f"is before {person.name}'s birth date "
                                        f"({person.birth_date.strftime('%d %b %Y')})."
                                    )
                                    break

                                # Age at marriage using simple year difference
                                age_at_marriage = marriage_date.year - person.birth_date.year
                                if (marriage_date.month, marriage_date.day) < (person.birth_date.month, person.birth_date.day):
                                    age_at_marriage -= 1

                                if age_at_marriage < 18:
                                    error = (
                                        f"{person.name} would be only {age_at_marriage} "
                                        f"years old at the time of marriage. Minimum age is 18."
                                    )
                                    break

                    except ValueError:
                        error = "Invalid marriage date format."
                # ────────────────────────────────────────────────

                if not error:
                    Marriage.objects.create(
                        spouse1=s1,
                        spouse2=s2,
                        marriage_date=date or None
                    )
                    return render(request, 'addRelationship.html', {
                        'success': True, 'males': males, 'females': females,
                    })

    return render(request, 'addRelationship.html', {
        'males': males, 'females': females, 'error': error,
    })


@login_required
def family_members(request):
    family  = get_user_family(request.user)
    persons = Person.objects.filter(family=family).select_related('father', 'mother', 'claimed_by')

    # Orphaned = no father, no mother, not joined_by_marriage
    orphans = persons.filter(
        father__isnull=True,
        mother__isnull=True,
        joined_by_marriage=False
    ).exclude(
        # Exclude true roots — males with no parents who ARE the tree starters
        # A root male has children, orphan males don't (optional filter)
    )

    total   = persons.count()
    males   = persons.filter(gender='M').count()
    females = persons.filter(gender='F').count()
    married = Marriage.objects.filter(spouse1__family=family).count()
    active  = persons.filter(claimed_by__isnull=False).count()

    return render(request, 'family_members.html', {
        'persons':  persons,
        'orphans':  orphans,
        'total':    total,
        'males':    males,
        'females':  females,
        'married':  married,
        'active':   active,
        'is_admin': is_family_admin(request.user, family),
        'family':   family,
    })

@login_required
def marriages_list(request):
    family    = get_user_family(request.user)
    marriages = Marriage.objects.filter(
        spouse1__family=family
    ).select_related('spouse1', 'spouse2', 'spouse1__claimed_by', 'spouse2__claimed_by')

    return render(request, 'marriages_list.html', {
        'marriages': marriages,
        'is_admin':  is_family_admin(request.user, family),
        'family':    family,
    })


def can_edit_marriage(user, marriage, family):
    """
    User can edit/delete a marriage if:
    - They are the family admin AND neither spouse has claimed their profile
    - OR they are one of the spouses (claimed the profile)
    """
    is_spouse = (
        marriage.spouse1.claimed_by == user or
        marriage.spouse2.claimed_by == user
    )
    if is_spouse:
        return True
    if is_family_admin(user, family):
        # Admin can only touch marriages where neither spouse is claimed
        if not marriage.spouse1.claimed_by and not marriage.spouse2.claimed_by:
            return True
    return False


@login_required
def edit_marriage(request, marriage_id):
    family   = get_user_family(request.user)
    marriage = get_object_or_404(
        Marriage,
        pk=marriage_id,
        spouse1__family=family
    )

    if not can_edit_marriage(request.user, marriage, family):
        return render(request, 'error.html', {
            'message': "You do not have permission to edit this marriage."
        })

    if request.method == 'POST':
        date = request.POST.get('marriage_date')
        marriage.marriage_date = date or None
        marriage.save()
        return redirect('marriages_list')

    return render(request, 'edit_marriage.html', {
        'marriage': marriage,
    })


@login_required
def delete_marriage(request, marriage_id):
    family   = get_user_family(request.user)
    marriage = get_object_or_404(
        Marriage,
        pk=marriage_id,
        spouse1__family=family
    )

    if not can_edit_marriage(request.user, marriage, family):
        return render(request, 'error.html', {
            'message': "You do not have permission to delete this marriage."
        })

    if request.method == 'POST':
        marriage.delete()
        return redirect('marriages_list')

    return render(request, 'confirm_delete_marriage.html', {
        'marriage': marriage,
    })



# ────────────────────────────────────────────────────To Export Data────────────────────────────────────────


@login_required
def export_excel(request):
    family  = get_user_family(request.user)
    persons = Person.objects.filter(family=family).select_related(
        'father', 'mother', 'claimed_by'
    ).order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{family.name} — Members"

    # ── Styles ──
    header_fill   = PatternFill("solid", fgColor="2C3E7A")
    male_fill     = PatternFill("solid", fgColor="DBE9FF")
    female_fill   = PatternFill("solid", fgColor="FCE7F3")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    title_font    = Font(bold=True, size=16, color="2C2040")
    bold_font     = Font(bold=True, size=10)
    normal_font   = Font(size=10)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")
    thin_border   = Border(
        left=Side(style='thin', color='E0D8F0'),
        right=Side(style='thin', color='E0D8F0'),
        top=Side(style='thin', color='E0D8F0'),
        bottom=Side(style='thin', color='E0D8F0'),
    )

    # ── Title row ──
    ws.merge_cells('A1:I1')
    ws['A1'] = f"{family.name} — Family Members"
    ws['A1'].font      = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 36

    # ── Sub-title ──
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Total members: {persons.count()}   |   Exported on: {__import__('datetime').date.today().strftime('%d %b %Y')}"
    ws['A2'].font      = Font(size=9, color="9A8FAA", italic=True)
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 20

    ws.append([])  # empty row

    # ── Headers ──
    headers = ['#', 'Name', 'Gender', 'Father', 'Mother', 'Birth Date', 'Joined By Marriage', 'Claimed By', 'Bio']
    ws.append(headers)
    header_row = ws.max_row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[header_row].height = 28

    # ── Data rows ──
    for idx, person in enumerate(persons, 1):
        row = [
            idx,
            person.name,
            'Male' if person.gender == 'M' else 'Female' if person.gender == 'F' else 'Unknown',
            person.father.name if person.father else '—',
            person.mother.name if person.mother else '—',
            person.birth_date.strftime('%d %b %Y') if person.birth_date else '—',
            'Yes' if person.joined_by_marriage else 'No',
            person.claimed_by.username if person.claimed_by else '—',
            person.bio or '—',
        ]
        ws.append(row)
        data_row = ws.max_row
        fill = male_fill if person.gender == 'M' else female_fill if person.gender == 'F' else None
        for col_num in range(1, len(headers) + 1):
            cell            = ws.cell(row=data_row, column=col_num)
            cell.font       = normal_font
            cell.alignment  = center_align if col_num in [1, 3, 6, 7] else left_align
            cell.border     = thin_border
            if fill:
                cell.fill = fill
        ws.row_dimensions[data_row].height = 22

    # ── Column widths ──
    col_widths = [5, 22, 10, 20, 20, 14, 18, 16, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ── Marriages sheet ──
    ws2 = wb.create_sheet(title="Marriages")
    marriages = Marriage.objects.filter(spouse1__family=family).select_related('spouse1', 'spouse2')

    ws2.merge_cells('A1:D1')
    ws2['A1'] = f"{family.name} — Marriages"
    ws2['A1'].font      = title_font
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 36
    ws2.append([])

    m_headers = ['#', 'Husband', 'Wife', 'Marriage Date']
    ws2.append(m_headers)
    m_header_row = ws2.max_row
    for col_num, _ in enumerate(m_headers, 1):
        cell = ws2.cell(row=m_header_row, column=col_num)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
    ws2.row_dimensions[m_header_row].height = 28

    for idx, m in enumerate(marriages, 1):
        husband = m.spouse1 if m.spouse1.gender == 'M' else m.spouse2
        wife    = m.spouse2 if m.spouse2.gender == 'F' else m.spouse1
        row = [
            idx,
            husband.name,
            wife.name,
            m.marriage_date.strftime('%d %b %Y') if m.marriage_date else '—',
        ]
        ws2.append(row)
        data_row = ws2.max_row
        for col_num in range(1, 5):
            cell           = ws2.cell(row=data_row, column=col_num)
            cell.font      = normal_font
            cell.alignment = center_align
            cell.border    = thin_border
        ws2.row_dimensions[data_row].height = 22

    for i, w in enumerate([5, 25, 25, 16], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Response ──
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{family.name}_family.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    family  = get_user_family(request.user)
    persons = Person.objects.filter(family=family).select_related(
        'father', 'mother', 'claimed_by'
    ).order_by('id')
    marriages = Marriage.objects.filter(
        spouse1__family=family
    ).select_related('spouse1', 'spouse2')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{family.name}_family.pdf"'

    doc    = SimpleDocTemplate(response, pagesize=A4,
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    # ── Title ──
    title_style = ParagraphStyle(
        'Title', parent=styles['Title'],
        fontSize=22, textColor=colors.HexColor('#2c2040'),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'Sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#9a8faa'),
        spaceAfter=16,
    )

    story.append(Paragraph(f"{family.name}", title_style))
    story.append(Paragraph(
        f"Family Members Report &nbsp;·&nbsp; "
        f"Total: {persons.count()} members &nbsp;·&nbsp; "
        f"Exported: {__import__('datetime').date.today().strftime('%d %b %Y')}",
        sub_style
    ))
    story.append(Spacer(1, 0.3*cm))

    # ── Members table ──
    story.append(Paragraph("Members", ParagraphStyle(
        'H2', parent=styles['Heading2'],
        fontSize=13, textColor=colors.HexColor('#2c3e7a'), spaceAfter=8
    )))

    m_data = [['#', 'Name', 'Gender', 'Father', 'Mother', 'Birth Date']]
    for idx, p in enumerate(persons, 1):
        m_data.append([
            str(idx),
            p.name,
            'M' if p.gender == 'M' else 'F' if p.gender == 'F' else '?',
            p.father.name if p.father else '—',
            p.mother.name if p.mother else '—',
            p.birth_date.strftime('%d %b %Y') if p.birth_date else '—',
        ])

    col_widths_pdf = [1*cm, 4*cm, 1.5*cm, 4*cm, 4*cm, 3*cm]
    t = Table(m_data, colWidths=col_widths_pdf, repeatRows=1)
    t.setStyle(TableStyle([
        # Header
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#2c3e7a')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 9),
        ('ALIGN',       (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING',(0,0),(-1,0), 8),
        ('TOPPADDING',  (0,0), (-1,0), 8),
        # Data rows
        ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,1), (-1,-1), 8),
        ('ALIGN',       (0,1), (0,-1),  'CENTER'),
        ('ALIGN',       (2,1), (2,-1),  'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1),
         [colors.white, colors.HexColor('#f8f6fc')]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e0d8f0')),
        ('TOPPADDING',  (0,1), (-1,-1), 5),
        ('BOTTOMPADDING',(0,1),(-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.8*cm))

    # ── Marriages table ──
    if marriages.exists():
        story.append(Paragraph("Marriages", ParagraphStyle(
            'H2b', parent=styles['Heading2'],
            fontSize=13, textColor=colors.HexColor('#7a2c5a'), spaceAfter=8
        )))

        mar_data = [['#', 'Husband', 'Wife', 'Marriage Date']]
        for idx, m in enumerate(marriages, 1):
            husband = m.spouse1 if m.spouse1.gender == 'M' else m.spouse2
            wife    = m.spouse2 if m.spouse2.gender == 'F' else m.spouse1
            mar_data.append([
                str(idx),
                husband.name,
                wife.name,
                m.marriage_date.strftime('%d %b %Y') if m.marriage_date else '—',
            ])

        t2 = Table(mar_data, colWidths=[1*cm, 5*cm, 5*cm, 3.5*cm], repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#7a2c5a')),
            ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
            ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,0), 9),
            ('ALIGN',       (0,0), (-1,0), 'CENTER'),
            ('BOTTOMPADDING',(0,0),(-1,0), 8),
            ('TOPPADDING',  (0,0), (-1,0), 8),
            ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,1), (-1,-1), 8),
            ('ALIGN',       (0,1), (0,-1),  'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.white, colors.HexColor('#fdf4f9')]),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e0d8f0')),
            ('TOPPADDING',  (0,1), (-1,-1), 5),
            ('BOTTOMPADDING',(0,1),(-1,-1), 5),
        ]))
        story.append(t2)

    doc.build(story)
    return response    



